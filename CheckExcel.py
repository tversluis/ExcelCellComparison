# This program will check the contents of an excel description column with a "correct" templates contents and add contents to the column to match.
# The data input next to the columns will be given zeros to indicate no data for that description. The data will then be placed into a master excel file.

import sys
from colorama import Fore, init, Back, Style
import openpyxl
import re

init(convert=True)

try:
    print("\n")
    #path = "F:\ANGIE\Quota new2015\Automotive Shops Variance 19-20.xlsx"
    wb_TEMPLATE = openpyxl.load_workbook("F:\ANGIE\Quota new2015\!Variance Reports\Automotive Shops Variance INPUT-NEW-DATA.xlsx".strip())
    # from the active attribute 
    sheet_TEMPLATE = wb_TEMPLATE.active
    missingCount = 0
    rowDifference = 43

    for i in range(rowDifference+3, rowDifference+41):
        TEMPLATE_Description_Cell = sheet_TEMPLATE.cell(row = i+missingCount, column = 1)
        obj_Description_Cell = sheet_TEMPLATE.cell(row = i-rowDifference, column = 1)
        
        if obj_Description_Cell.value != TEMPLATE_Description_Cell.value:
            sheet_TEMPLATE.insert_rows(i-rowDifference)
            cell_value1 = sheet_TEMPLATE.cell(row = i-rowDifference, column = 1)
            cell_value2 = sheet_TEMPLATE.cell(row = i-rowDifference, column = 2)
            cell_value3 = sheet_TEMPLATE.cell(row = i-rowDifference, column = 3)
            cell_value4 = sheet_TEMPLATE.cell(row = i-rowDifference, column = 4)
            cell_value1.value = TEMPLATE_Description_Cell.value
            cell_value2.value = 0
            cell_value3.value = 0
            cell_value4.value = 0
            missingCount= missingCount+1

    wb_TEMPLATE.save(filename = "F:\ANGIE\Quota new2015\!Variance Reports\Automotive Shops Variance OUTPUT-NEW-DATA.xlsx")

    referenceCell1 = sheet_TEMPLATE.cell(row=3, column=1).value
    referenceCell2 = sheet_TEMPLATE.cell(row=3, column=2).value
    referenceCell3 = sheet_TEMPLATE.cell(row=3, column=3).value
    referenceCell4 = sheet_TEMPLATE.cell(row=3, column=4).value
    print("REFERENCED CELLS \"" + str(referenceCell1) + ": " + str(referenceCell2) + ", " + str(referenceCell3) + ", " + str(referenceCell4) + "\"")

    for j in range(1, 4):
        for k in range(3, 43):
            sheet_TEMPLATE.cell(row = k, column = j).value=None
    

except Exception as e:
    print(e)
    print (Fore.RED + "Error : The file is not found")
print(Fore.GREEN + "###################### Success! Excel file has been read/written. ##############################")
